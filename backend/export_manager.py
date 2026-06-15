"""
Export Manager for Timetable Generator.

Handles exporting timetables to various formats (PDF, Excel, CSV, iCal).
"""

import csv
import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional

from .db_config import get_db_connection

class ExportManager:
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = export_dir
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
    
    def _format_time(self, time_val):
        """Helper function to format time values"""
        if time_val is None:
            return ""
        time_str = str(time_val)
        if ':' in time_str:
            parts = time_str.split(':')
            hours = int(parts[0])
            mins = parts[1] if len(parts) > 1 else '00'
            return f"{hours:02d}:{mins[:2]}"
        return time_str
    
    def export_to_csv(self, section_id: Optional[int] = None, teacher_id: Optional[int] = None, 
                     room_id: Optional[int] = None, version_id: Optional[int] = None, 
                     user_id: Optional[int] = None) -> str:
        """Export timetable to CSV format"""
        slots = self._load_timetable_data(section_id, teacher_id, room_id, version_id)
        
        filename = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.export_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Day', 'Start Time', 'End Time', 'Subject', 'Teacher', 'Section', 'Room', 'Room Type']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for slot in slots:
                writer.writerow({
                    'Day': slot['day_of_week'],
                    'Start Time': str(slot['start_time']),
                    'End Time': str(slot['end_time']),
                    'Subject': slot.get('subject_name', ''),
                    'Teacher': slot.get('teacher_name', ''),
                    'Section': slot.get('section_name', ''),
                    'Room': slot.get('room_name', ''),
                    'Room Type': slot.get('room_type', '')
                })
        
        self._log_export('CSV', section_id, teacher_id, room_id, filepath, user_id)
        return filepath
    
    def export_to_excel(self, section_id: Optional[int] = None, teacher_id: Optional[int] = None,
                       room_id: Optional[int] = None, version_id: Optional[int] = None, 
                       user_id: Optional[int] = None) -> str:
        """Export timetable to Excel format - Grid format (same as web page)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        slots = self._load_timetable_data(section_id, teacher_id, room_id, version_id)
        
        filename = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        # Get section name for title
        section_name = "All Classes"
        if section_id:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            try:
                cursor.execute("""
                    SELECT p.name as program_name, s.section_name, s.year
                    FROM sections s
                    JOIN programs p ON s.program_id = p.program_id
                    WHERE s.section_id = %s
                """, (section_id,))
                section = cursor.fetchone()
                if section:
                    section_name = f"{section['program_name']} - {section['section_name']} (Year {section['year']})"
            finally:
                cursor.close()
                conn.close()
        
        # Organize slots into grid format
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        
        # Get all unique time slots
        time_slots_set = set()
        day_slot_map = {}  # (day, timeslot_key) -> slot_data
        
        for slot in slots:
            day = slot['day_of_week']
            if day in day_order:
                start_time = self._format_time(slot['start_time'])
                end_time = self._format_time(slot['end_time'])
                time_key = f"{start_time}-{end_time}"
                time_slots_set.add(time_key)
                day_slot_map[(day, time_key)] = slot
        
        # Sort time slots
        time_slots = sorted(time_slots_set, key=lambda x: (
            int(x.split('-')[0].split(':')[0]) * 60 + int(x.split('-')[0].split(':')[1])
        ))
        
        if not time_slots:
            # No slots, create empty workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timetable"
            ws.append(["No timetable data available"])
            wb.save(filepath)
            return filepath
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:' + openpyxl.utils.get_column_letter(len(time_slots) + 1) + '1')
        ws['A1'] = f"Timetable - {section_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Column headers (time slots)
        ws.append(["Day / Time"] + time_slots)  # Row 2
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for i, ts in enumerate(time_slots):
            col_letter = openpyxl.utils.get_column_letter(i + 2)
            ws.column_dimensions[col_letter].width = 20
        
        # Add data rows for each day
        row_num = 3
        for day in day_order:
            row_data = [day]
            for time_key in time_slots:
                slot = day_slot_map.get((day, time_key))
                if slot:
                    # Format: Subject\nTeacher\nRoom
                    content = f"{slot.get('subject_name', '')}\n{slot.get('teacher_name', '')}\n{slot.get('room_name', '')}"
                    row_data.append(content)
                else:
                    row_data.append("-")
            
            ws.append(row_data)
            
            # Style day column
            ws.cell(row=row_num, column=1).fill = day_fill
            ws.cell(row=row_num, column=1).font = day_font
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=1).border = thin_border
            
            # Style time slot cells
            for col_num in range(2, len(time_slots) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                # Add light fill for occupied cells
                if ws.cell(row=row_num, column=col_num).value and ws.cell(row=row_num, column=col_num).value != "-":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            row_num += 1
        
        # Freeze top row and first column
        ws.freeze_panes = 'B3'
        
        wb.save(filepath)
        self._log_export('EXCEL', section_id, teacher_id, room_id, filepath, user_id)
        return filepath

    def export_multiple_sections_to_excel(
        self,
        section_ids: List[int],
        user_id: Optional[int] = None
    ) -> str:
        """
        Export multiple section timetables into a single Excel file.

        Each section gets its own sheet, using the same grid layout as the web view.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not section_ids:
            raise ValueError("At least one section_id is required for multiple-section export")

        filename = f"timetables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        wb = openpyxl.Workbook()
        # Remove the default sheet; we'll create specific ones below
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Common styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Pre-load section metadata
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        section_meta = {}
        try:
            placeholders = ",".join(["%s"] * len(section_ids))
            cursor.execute(
                f"""
                SELECT s.section_id, p.name as program_name, s.section_name, s.year
                FROM sections s
                JOIN programs p ON s.program_id = p.program_id
                WHERE s.section_id IN ({placeholders})
                """,
                tuple(section_ids),
            )
            for row in cursor.fetchall():
                section_meta[row["section_id"]] = row
        finally:
            cursor.close()
            conn.close()

        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

        for section_id in section_ids:
            # Load data for this section
            slots = self._load_timetable_data(section_id=section_id)

            meta = section_meta.get(section_id)
            if meta:
                section_name = f"{meta['program_name']} - {meta['section_name']} (Year {meta['year']})"
                sheet_title = f"{meta['section_name']}".strip()[:31] or f"Class_{section_id}"
            else:
                section_name = f"Section {section_id}"
                sheet_title = f"Class_{section_id}"

            ws = wb.create_sheet(title=sheet_title)

            # Collect timeslots for this section
            time_slots_set = set()
            day_slot_map = {}
            for slot in slots:
                day = slot['day_of_week']
                if day in day_order:
                    start_time = self._format_time(slot['start_time'])
                    end_time = self._format_time(slot['end_time'])
                    time_key = f"{start_time}-{end_time}"
                    time_slots_set.add(time_key)
                    day_slot_map[(day, time_key)] = slot

            time_slots = sorted(
                time_slots_set,
                key=lambda x: int(x.split('-')[0].split(':')[0]) * 60 + int(x.split('-')[0].split(':')[1]),
            )

            # If no data for this section, put a simple message and continue
            if not time_slots:
                ws.append([f"No timetable data available for {section_name}"])
                continue

            # Title
            ws.merge_cells('A1:' + openpyxl.utils.get_column_letter(len(time_slots) + 1) + '1')
            ws['A1'] = f"Timetable - {section_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")

            # Header row with time slots
            ws.append(["Day / Time"] + list(time_slots))
            for cell in ws[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Column widths
            ws.column_dimensions['A'].width = 15
            for i, _ in enumerate(time_slots):
                col_letter = openpyxl.utils.get_column_letter(i + 2)
                ws.column_dimensions[col_letter].width = 20

            # Data rows
            row_num = 3
            for day in day_order:
                row_data = [day]
                for time_key in time_slots:
                    slot = day_slot_map.get((day, time_key))
                    if slot:
                        content = f"{slot.get('subject_name', '')}\n{slot.get('teacher_name', '')}\n{slot.get('room_name', '')}"
                        row_data.append(content)
                    else:
                        row_data.append("-")

                ws.append(row_data)

                ws.cell(row=row_num, column=1).fill = day_fill
                ws.cell(row=row_num, column=1).font = day_font
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_num, column=1).border = thin_border

                for col_num in range(2, len(time_slots) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                    if cell.value and cell.value != "-":
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                row_num += 1

            ws.freeze_panes = 'B3'

        wb.save(filepath)

        # Log one export record with scope FULL to avoid over-logging
        self._log_export('EXCEL_MULTI', None, None, None, filepath, user_id)
        return filepath
    
    def export_to_ical(self, section_id: Optional[int] = None, teacher_id: Optional[int] = None,
                      room_id: Optional[int] = None, version_id: Optional[int] = None, 
                      user_id: Optional[int] = None) -> str:
        """Export timetable to iCal format"""
        slots = self._load_timetable_data(section_id, teacher_id, room_id, version_id)
        
        filename = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.ics"
        filepath = os.path.join(self.export_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("BEGIN:VCALENDAR\n")
            f.write("VERSION:2.0\n")
            f.write("PRODID:-//Timetable Generator//EN\n")
            f.write("CALSCALE:GREGORIAN\n")
            f.write("METHOD:PUBLISH\n")
            
            # Get current year for dates
            current_year = datetime.now().year
            
            for slot in slots:
                # Convert day name to date (simplified - using first occurrence in current year)
                day_name = slot['day_of_week']
                day_map = {
                    'Monday': 0, 'Tuesday': 1, 'Wednesday': 2,
                    'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6
                }
                
                # Find first Monday of year as reference
                jan1 = datetime(current_year, 1, 1)
                days_offset = (day_map[day_name] - jan1.weekday()) % 7
                event_date = jan1 + timedelta(days=days_offset)
                
                start_time = datetime.strptime(str(slot['start_time']), '%H:%M:%S').time()
                end_time = datetime.strptime(str(slot['end_time']), '%H:%M:%S').time()
                
                dtstart = datetime.combine(event_date, start_time)
                dtend = datetime.combine(event_date, end_time)
                
                # Format for iCal (UTC)
                dtstart_str = dtstart.strftime('%Y%m%dT%H%M%S')
                dtend_str = dtend.strftime('%Y%m%dT%H%M%S')
                
                summary = f"{slot.get('subject_name', 'Class')} - {slot.get('section_name', '')}"
                description = f"Teacher: {slot.get('teacher_name', '')}\\nRoom: {slot.get('room_name', '')}"
                location = slot.get('room_name', '')
                
                f.write("BEGIN:VEVENT\n")
                f.write(f"DTSTART:{dtstart_str}\n")
                f.write(f"DTEND:{dtend_str}\n")
                f.write(f"SUMMARY:{summary}\n")
                f.write(f"DESCRIPTION:{description}\n")
                f.write(f"LOCATION:{location}\n")
                f.write("RRULE:FREQ=WEEKLY;COUNT=52\n")  # Repeat weekly for a year
                f.write("END:VEVENT\n")
            
            f.write("END:VCALENDAR\n")
        
        self._log_export('ICAL', section_id, teacher_id, room_id, filepath, user_id)
        return filepath
    
    def export_to_pdf(self, section_id: Optional[int] = None, teacher_id: Optional[int] = None,
                     room_id: Optional[int] = None, version_id: Optional[int] = None, 
                     user_id: Optional[int] = None) -> str:
        """Export timetable to PDF format - Grid format (same as web page)"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        slots = self._load_timetable_data(section_id, teacher_id, room_id, version_id)
        
        # Get section name for title
        section_name = "All Classes"
        if section_id:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            try:
                cursor.execute("""
                    SELECT p.name as program_name, s.section_name, s.year
                    FROM sections s
                    JOIN programs p ON s.program_id = p.program_id
                    WHERE s.section_id = %s
                """, (section_id,))
                section = cursor.fetchone()
                if section:
                    section_name = f"{section['program_name']} - {section['section_name']} (Year {section['year']})"
            finally:
                cursor.close()
                conn.close()
        
        filename = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.export_dir, filename)
        
        # Use landscape for better grid view
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=0.2*inch, rightMargin=0.2*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,  # Center
            spaceAfter=12
        )
        
        # Title
        title = Paragraph(f"Timetable - {section_name}", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2*inch))
        
        # Organize slots into grid format
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        
        # Get all unique time slots
        time_slots_set = set()
        day_slot_map = {}  # (day, timeslot_key) -> slot_data
        
        for slot in slots:
            day = slot['day_of_week']
            if day in day_order:
                start_time = self._format_time(slot['start_time'])
                end_time = self._format_time(slot['end_time'])
                time_key = f"{start_time}-{end_time}"
                time_slots_set.add(time_key)
                day_slot_map[(day, time_key)] = slot
        
        # Sort time slots
        def time_to_minutes(time_str):
            try:
                parts = time_str.split('-')[0].split(':')
                return int(parts[0]) * 60 + int(parts[1])
            except:
                return 0
        
        time_slots = sorted(time_slots_set, key=time_to_minutes)
        
        if not time_slots:
            story.append(Paragraph("No timetable data available"))
            doc.build(story)
            return filepath
        
        # Build the grid table
        # Header row: Day/Time + time slots
        header_row = ['Day / Time'] + time_slots
        
        # Data rows: each day with its slots
        table_data = [header_row]
        
        for day in day_order:
            row = [day]
            for time_key in time_slots:
                slot = day_slot_map.get((day, time_key))
                if slot:
                    # Format: Subject\nTeacher\nRoom
                    content = f"{slot.get('subject_name', '')}<br/>{slot.get('teacher_name', '')}<br/>{slot.get('room_name', '')}"
                    row.append(content)
                else:
                    row.append("-")
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        
        # Calculate column widths
        num_cols = len(time_slots) + 1
        col_widths = [0.8*inch] + [1.5*inch] * len(time_slots)
        
        # Style the table
        style = TableStyle([
            # Header row style
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.21, 0.27, 0.36)),  # #366092
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # First column (day names) style
            ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.27, 0.33, 0.44)),  # #4472C4
            ('TEXTCOLOR', (0, 1), (0, -1), colors.whitesmoke),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Cell padding
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ])
        
        # Add alternating row colors for data rows
        for i in range(1, len(day_order) + 1):
            if i % 2 == 0:
                style.add('BACKGROUND', (1, i), (-1, i), colors.Color(0.95, 0.95, 0.95))
            else:
                style.add('BACKGROUND', (1, i), (-1, i), colors.white)
            
            # Highlight occupied cells
            for j in range(1, num_cols):
                cell_text = table_data[i][j]
                if cell_text != "-":
                    style.add('BACKGROUND', (j, i), (j, i), colors.Color(0.88, 0.94, 0.85))  # Light green
        
        table.setStyle(style)
        story.append(table)
        
        doc.build(story)
        self._log_export('PDF', section_id, teacher_id, room_id, filepath, user_id)
        return filepath
    
    def _load_timetable_data(self, section_id: Optional[int] = None, 
                            teacher_id: Optional[int] = None,
                            room_id: Optional[int] = None,
                            version_id: Optional[int] = None) -> List[Dict]:
        """Load timetable data based on filters"""
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            if version_id:
                query = """
                    SELECT ts.*, t.day_of_week, t.start_time, t.end_time,
                           s.name as subject_name,
                           te.name as teacher_name,
                           sec.section_name,
                           r.name as room_name, r.type as room_type
                    FROM timetable_version_slots ts
                    JOIN timeslots t ON ts.timeslot_id = t.timeslot_id
                    LEFT JOIN subjects s ON ts.subject_id = s.subject_id
                    LEFT JOIN teachers te ON ts.teacher_id = te.teacher_id
                    LEFT JOIN sections sec ON ts.section_id = sec.section_id
                    LEFT JOIN rooms r ON ts.room_id = r.room_id
                    WHERE ts.version_id = %s
                """
                params = [version_id]
            else:
                query = """
                    SELECT ts.*, t.day_of_week, t.start_time, t.end_time,
                           s.name as subject_name,
                           te.name as teacher_name,
                           sec.section_name,
                           r.name as room_name, r.type as room_type
                    FROM timetable_slots ts
                    JOIN timeslots t ON ts.timeslot_id = t.timeslot_id
                    LEFT JOIN subjects s ON ts.subject_id = s.subject_id
                    LEFT JOIN teachers te ON ts.teacher_id = te.teacher_id
                    LEFT JOIN sections sec ON ts.section_id = sec.section_id
                    LEFT JOIN rooms r ON ts.room_id = r.room_id
                    WHERE 1=1
                """
                params = []
            
            if section_id:
                query += " AND ts.section_id = %s"
                params.append(section_id)
            
            if teacher_id:
                query += " AND ts.teacher_id = %s"
                params.append(teacher_id)
            
            if room_id:
                query += " AND ts.room_id = %s"
                params.append(room_id)
            
            query += " ORDER BY t.day_of_week, t.start_time"
            
            cursor.execute(query, tuple(params))
            return cursor.fetchall()
        finally:
            cursor.close()
            conn.close()
    
    def _log_export(self, export_type: str, section_id: Optional[int], 
                   teacher_id: Optional[int], room_id: Optional[int], filepath: str, 
                   user_id: Optional[int]):
        """Log export to database"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Determine scope
            if section_id:
                scope = 'SECTION'
                scope_id = section_id
            elif teacher_id:
                scope = 'TEACHER'
                scope_id = teacher_id
            elif room_id:
                scope = 'ROOM'
                scope_id = room_id
            else:
                scope = 'FULL'
                scope_id = None
            
            cursor.execute("""
                INSERT INTO export_history 
                (user_id, export_type, export_scope, scope_id, file_path, created_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
            """, (user_id, export_type, scope, scope_id, filepath))
            
            conn.commit()
        except Exception as e:
            print(f"Error logging export: {e}")
        finally:
            cursor.close()
            conn.close()
